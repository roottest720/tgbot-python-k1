import telebot
import string
import random
import openpyxl
import requests

BOT_TOKEN = "6469470826:AAEl8-HcgfzcrlOW5_qft3ncxtLDQ7IALaE"

channel_id = "-1001867856337"
join_link = "https://telegram.dog/+u3MvJUAwUVI2YWQ9"
base_link = "https://telegram.dog/kiara_video_send_bot?start="
urlshortx_api = "https://urlshortx.com/api?api=e7b8d7ff3408755f4a812abece62044ee29c82f1&url="
admin_id = "5113588348"

users_excel_file = "users.xlsx"
workbook = openpyxl.load_workbook(users_excel_file)
worksheet = workbook.active

videos_excel_file = "videos.xlsx"
workbook1 = openpyxl.load_workbook(videos_excel_file)
worksheet1 = workbook1.active

tokens_excel_file = "tokens.xlsx"
workbook2 = openpyxl.load_workbook(tokens_excel_file)
worksheet2 = workbook2.active

bot = telebot.TeleBot(BOT_TOKEN)


def randomStringGenerator(n=10):
    """Random string generator ( Default length 10 )"""
    return ''.join(random.choices(string.ascii_uppercase +string.digits +string.ascii_lowercase, k=n))

def generateToken():
    token = "token_" + randomStringGenerator(20)
    new_row_data = [token, "unused"]
    worksheet2.append(new_row_data)
    workbook2.save(tokens_excel_file)
    return token

def verfiyToken(token):
    target_data = token
    row_number = None

    for row_num, row in enumerate(worksheet2.iter_rows(values_only=True), start=1):
        cell_value = row[0]  # Assuming you are searching in the first column (column index 0)
        if cell_value == target_data:
            row_number = row_num
            break

    if row_number is not None:
        column_number = 2
        cell_value = worksheet2.cell(row=row_number, column=column_number).value
        if cell_value is not None:
            if cell_value == "unused":
                worksheet2.cell(row=row_number, column=column_number, value="used")
                workbook2.save(tokens_excel_file)
                return True
            else:
                return False
        else:
            return False
    else:
        return False

def shortLink(token):
    token = base_link + token
    response = requests.get(urlshortx_api + token)
    if response.status_code == 200:
        data = response.json()
        return data["shortenedUrl"]
    else:
        return 1

def addUser(user_id):
    is_empty = True

    for row in worksheet.iter_rows():
        for cell in row:
            if cell.value is not None:
                is_empty = False
                break

        if not is_empty:
            break

    if is_empty:
        new_row_data = [user_id, 25]
        worksheet.append(new_row_data)
        workbook.save(users_excel_file)
    else:
        target_data = user_id
        found = False
        for row in worksheet.iter_rows():
            for cell in row:
                if cell.value == target_data:
                    found = True
                    break

            if found:
                break
        if found == False:
            new_row_data = [user_id, 25]
            worksheet.append(new_row_data)
            workbook.save(users_excel_file)

def getCredit(user_id):
    # user_id = str(user_id)
    target_data = user_id
    row_number = None

    for row_num, row in enumerate(worksheet.iter_rows(values_only=True), start=1):
        cell_value = row[0]  # Assuming you are searching in the first column (column index 0)
        if cell_value == target_data:
            row_number = row_num
            break

    if row_number is not None:
        column_number = 2
        cell_value = worksheet.cell(row=row_number, column=column_number).value
        if cell_value is not None:
            return cell_value
        else:
            return 0
    else:
        return 0
def decreaseCredit(user_id):
    target_data = user_id
    row_number = None

    for row_num, row in enumerate(worksheet.iter_rows(values_only=True), start=1):
        cell_value = row[0]  # Assuming you are searching in the first column (column index 0)
        if cell_value == target_data:
            row_number = row_num
            break

    if row_number is not None:
        column_number = 2
        cell_value = worksheet.cell(row=row_number, column=column_number).value
        if cell_value is not None:
            credit = cell_value
            if credit > 0:
                decreased_credit = credit-1;
                worksheet.cell(row=row_number, column=column_number, value=decreased_credit)
                workbook.save(users_excel_file)
                return decreased_credit
            else:
                return 0
        else:
            return 0
    else:
        return 0
    
def addCredit(user_id,creditToAdd=25):
    target_data = user_id
    row_number = None

    for row_num, row in enumerate(worksheet.iter_rows(values_only=True), start=1):
        cell_value = row[0]  # Assuming you are searching in the first column (column index 0)
        if cell_value == target_data:
            row_number = row_num
            break

    if row_number is not None:
        column_number = 2
        cell_value = worksheet.cell(row=row_number, column=column_number).value
        if cell_value is not None:
            credit = cell_value + creditToAdd
            worksheet.cell(row=row_number, column=column_number, value=credit)
            workbook.save(users_excel_file)
            return True
        else:
            return False
    else:
        return False
print('started')

@bot.message_handler(commands=['start'])
def send_welcome(message):
    addUser(message.from_user.id)
    userStatus = bot.get_chat_member(channel_id,message.from_user.id).status
    
    if len(message.text.split()) == 2:
        if (userStatus == "creator") or (userStatus == "member") or (userStatus == "administrator"):
            uid = message.text.split()[1]
            if uid.split('_')[0] == "token":
                if verfiyToken(uid):
                    if addCredit(message.from_user.id,25):
                        bot.reply_to(message,"25 credit added.\nTotal credits: " + str(getCredit(message.from_user.id)))
                else:
                    bot.reply_to(message,"Invalid or used token")
            else:
                target_data = uid
                row_number = None

                for row_num, row in enumerate(worksheet1.iter_rows(values_only=True), start=1):
                    cell_value = row[0]  # Assuming you are searching in the first column (column index 0)
                    if cell_value == target_data:
                        row_number = row_num
                        break

                if row_number is not None:
                    column_number = 2
                    cell_value = worksheet1.cell(row=row_number, column=column_number).value
                    if cell_value is not None:
                        credit = getCredit(message.from_user.id)
                        if credit <= 0:
                            token = generateToken()
                            link = shortLink(token)
                            if link == 1:
                                bot.reply_to(message,"Something went wrong! Please contact admin.")
                            else:
                                button_1 = telebot.types.InlineKeyboardButton('Add 25 credits', url=link)
                                keyboard = telebot.types.InlineKeyboardMarkup()
                                keyboard.add(button_1)
                                bot.reply_to(message,"Your credit is 0.",reply_markup=keyboard)
                        else:
                            bot.send_video(message.from_user.id,cell_value)
                            dCredit = decreaseCredit(message.from_user.id)
                            if dCredit == 0:
                                bot.send_message(message.from_user.id,"Now, you have 0 credit.")
                            elif dCredit % 5 == 0:
                                bot.send_message(message.from_user.id,f"{dCredit} credits left.")
                    else:
                        bot.reply_to(message,"Invalid link!")
                else:
                    bot.reply_to(message,"Invalid link!")
        else:
            button_1 = telebot.types.InlineKeyboardButton('Join', url=join_link)
            button_2 = telebot.types.InlineKeyboardButton('Check', url=base_link + message.text.split()[1])

            keyboard = telebot.types.InlineKeyboardMarkup()
            keyboard.add(button_1,button_2)
            # keyboard.add(button_2)

            bot.reply_to(message, "Hey "+ message.from_user.first_name +", \nYou have to <b>join our channel</b> to access this content.\nPlease click on <b>\"Join\"</b> button and after joining click <b>\"Check\"</b> button.",parse_mode='HTML', reply_markup=keyboard)

    else:
        button_1 = telebot.types.InlineKeyboardButton('Join', url=join_link)
        keyboard = telebot.types.InlineKeyboardMarkup()
        keyboard.add(button_1)
        bot.reply_to(message, "Hey "+ message.from_user.first_name +", \nYou have <b>" + str(getCredit(message.from_user.id)) + " credits</b>\n\nYou can use these credits to get videos from given channel.",parse_mode='HTML',reply_markup=keyboard)    

@bot.message_handler(commands=['me'])
def about_me(message):
    bot.reply_to(message, bot.get_me())

@bot.message_handler(commands=['help'])
def about_me(message):
    bot.reply_to(message, "Hello, this bot is file send bot.\n\nIts work on credit system.\nEvery time you request a video you loss 1 credit.\n\nThese are some useful commands of this bot:\n/add_credits ( To add more credits )\n/my_credits ( To check how many credits you have )")

@bot.message_handler(commands=['add_credits'])
def about_me(message):
    token = generateToken()
    link = shortLink(token)
    button_1 = telebot.types.InlineKeyboardButton('Add 25 credits', url=link)
    keyboard = telebot.types.InlineKeyboardMarkup()
    keyboard.add(button_1)
    bot.reply_to(message,"You have " + str(getCredit(message.from_user.id)) + " credits left.",reply_markup=keyboard)

@bot.message_handler(commands=['my_credits'])
def about_me(message):
    bot.reply_to(message,"You have " + str(getCredit(message.from_user.id)) + " credits left.")

@bot.message_handler(commands=['total_users'])
def about_me(message):
    if admin_id == str(message.from_user.id):
        num_rows_with_data = 0
        for row in worksheet.iter_rows(values_only=True):
            if any(cell is not None for cell in row):
                num_rows_with_data += 1
        bot.reply_to(message,"Total users: " + str(num_rows_with_data))

# reply for text 
@bot.message_handler(func=lambda msg: True)
def echo_all(message):
    bot.reply_to(message,"I am file sending bot and I only reply to file links.")

# Handles all sent vedio
@bot.message_handler(content_types=['video'])
def handle_videos(message):
    if admin_id == str(message.from_user.id):
        uid = randomStringGenerator(15)
        file_id = message.video.file_id
        new_row_data = [uid, file_id]
        worksheet1.append(new_row_data)
        workbook1.save(videos_excel_file)
        bot.reply_to(message,"Link: " + base_link + uid)

bot.infinity_polling()
